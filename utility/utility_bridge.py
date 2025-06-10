import webbrowser
import sys
import _thread
import os
import subprocess
from conf.conf_bridge import CONFIGURATION as conf
import time
import traceback
import textwrap
import psutil
import mysql.connector
import shutil
from win32com import client as win32client
from datetime import date,datetime
from pathlib import Path
import threading
import pythoncom
import openpyxl
from openpyxl.styles import Alignment, Font
from win32com.client import DispatchEx
import re
import PyPDF2
'''===============================Original function========================================'''
def open_folder(dir):
    webbrowser.open(dir)

def open_link_with_edge(link):
    edge_address = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
    def open_link():
        subprocess.call([edge_address, link])
    _thread.start_new_thread(open_link, ())

def file_exists(file_dir):
    return os.path.exists(file_dir)

def is_float(input):
    try:
        float(input)
        return True
    except:
        return False

def is_str(input):
    try:
        str(input)
        return True
    except:
        return False

def create_directory(directory):
    if os.path.exists(directory) == False:
        try:
            os.makedirs(directory)
        except Exception as e:
            print(f"Failed to create directory {directory}. Reason: {e}")


'''================================New function================================='''



def open_pdf_thread(pdf_path):
    try:
        _thread.start_new_thread(open_pdf, (pdf_path,))
    except:
        traceback.print_exc()

def open_pdf(pdf_path):
    try:
        adobe_address=conf["adobe_address"]
        subprocess.call([adobe_address, pdf_path])
    except:
        traceback.print_exc()

def kill_adobe():
    for proc in psutil.process_iter():
        if proc.name() == "Acrobat.exe":
            proc.kill()

def get_new_quotation_number():
    mydb = mysql.connector.connect(host="192.168.1.199", user="bridge", password="PcE$yD2023", database="bridge")
    mycursor = mydb.cursor()
    mycursor.execute("SELECT MAX(quotation_number) AS max_value FROM bridge.projects")
    current_quo_number = str(mycursor.fetchall()[0][0])
    if not current_quo_number.startswith(date.today().strftime("%y%m000")[1:4]):
        new_quotation = date.today().strftime("%y%m000")[1:] + "AA"
    else:
        quotation_letter = current_quo_number[6:8][0] + chr(ord(current_quo_number[6:8][1]) + 1) if current_quo_number[6:8][1] != "Z" else chr(ord(current_quo_number[6:8][0]) + 1) + "A"
        new_quotation = current_quo_number[:6] + quotation_letter
    return new_quotation







def create_new_project(quotation,folder_name):
    try:
        accounting_dir = os.path.join(conf["accounting_dir"], quotation)
        folder_path = folder_name
        calculation_sheet = conf["calculation_sheet"]
        check_or_create_folder(accounting_dir)
        check_or_create_folder(folder_path)
        os.mkdir(os.path.join(folder_path, "External"))
        os.mkdir(os.path.join(folder_path, "Photos"))
        os.mkdir(os.path.join(folder_path, "Plot"))
        os.mkdir(os.path.join(folder_path, "SS"))
        shutil.copyfile(os.path.join(conf["resource_dir"], "xlsx", calculation_sheet),os.path.join(folder_path, calculation_sheet))
        shortcut_dir = os.path.join(folder_path, "Database Shortcut.lnk")
        shortcut_working_dir = conf["accounting_dir"]
        shell = win32client.Dispatch("WScript.Shell")
        shortcut = shell.CreateShortcut(shortcut_dir)
        shortcut.Targetpath = accounting_dir
        shortcut.WorkingDirectory = shortcut_working_dir
        shortcut.save()
        create_new_project_database(quotation,folder_name)

    except:
        traceback.print_exc()

def delete_project_database(quo_num):
    try:
        mydb = mysql.connector.connect(host="192.168.1.199", user="bridge", password="PcE$yD2023", database="bridge")
        mycursor = mydb.cursor()
        table_names =['copilot_tags','copilot_sizes','man_power','remittance_upload','supplier_upload','bill_upload',
                      'acceptance_upload','status_colour','invoices','fee_acceptance','supplier_fee','bills',
                      'remittances','emails','fee_item','fee','scope','major_stage',
                      'user_history','drawings','apartments','rooms','levels','projects']
        for i in range(2):
            for table_name in table_names:
                try:
                    mycursor.execute(f"DELETE FROM bridge.{table_name} WHERE quotation_number = %s", (quo_num,))
                except:
                    pass
        mydb.commit()
    except:
        traceback.print_exc()
    finally:
        mycursor.close()
        mydb.close()
    try:
        mydb = mysql.connector.connect(host="192.168.1.199", user="bridge", password="PcE$yD2023", database="bridge")
        mycursor = mydb.cursor()
        mycursor.execute(f"DELETE FROM bridge.projects WHERE quotation_number = %s", (quo_num,))
        mydb.commit()
    except:
        traceback.print_exc()
    finally:
        mycursor.close()
        mydb.close()



def create_new_project_database(quotation,folder_name):
    folder=Path(folder_name).name
    pro_name='-'.join(folder.split('-')[1:])
    mydb = mysql.connector.connect(host="192.168.1.199", user="bridge", password="PcE$yD2023", database="bridge")
    mycursor = mydb.cursor()
    '''bridge.projects'''
    date_today = str(datetime.now().strftime('%Y-%m-%d'))
    asana_note=f'Email sent from ??? on {date_today}, with Architectural Plans.\nThe project is a residential develop and consists of:\n-      ??? levels of Basement car park, approximately ??? car spots.\n-      Ground level with a child care\n-      Building A ground to level ???, with Residential Apartments, approximately ??? apartments\n-      Building B ground to level ???, with Residential Apartments, approximately ??? apartments'
    mycursor.execute("INSERT INTO bridge.projects VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                     (quotation,None,pro_name,'Set Up',
                      'Client',None,None,None,0,0,0,
                      0,0,0,0,0,0,
                      None,None,None,None,None,asana_note,None,None,
                      folder,None,None,0,0,0))
    '''bridge.levels'''
    levels_dict={0:'Tenancy Level',1:'',2:'',3:'',4:''}
    for index,level_name in levels_dict.items():
        mycursor.execute("INSERT INTO bridge.levels (quotation_number, levels,spaces,area,row_index)"
                         "VALUES (%s, %s, %s, %s, %s)",
                         (quotation,level_name,None,None,index))
    '''bridge.rooms'''
    for i in range(8):
        mycursor.execute("INSERT INTO bridge.rooms VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                         (quotation, None, None, None, None, None, None, None, None, i))
    '''bridge.apartments'''
    for i in range(20):
        mycursor.execute("INSERT INTO bridge.apartments VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                         (quotation, None, None, None, None, None, None, None, None, i))
    '''bridge.drawings'''
    drawing_dict={0:('M-000','Cover Sheet','A'),
                  1:('M-100','Tenancy Level Layout','A'),
                  2:('M-101','Roof Layout','A'),
                  3:(None,None,None),
                  4:(None,None,None),
                  5:(None,None,None),
                  6:(None,None,None),
                  7:(None,None,None),
                  8:(None,None,None),
                  9:(None,None,None),
                  10:(None,None,None),
                  11:(None,None,None)}
    for index,content in drawing_dict.items():
        drawing_num=content[0]
        drawing_name=content[1]
        drawing_rev=content[2]
        mycursor.execute("INSERT INTO bridge.drawings VALUES (%s, %s, %s, %s, %s)",
                         (quotation, drawing_num, drawing_name, drawing_rev, index))
    '''bridge.fee'''
    installation_program='Week 1: site induction, site inspection and site measure, coordination.\nWeek 2-3: Order ductwork, VCDs, material and arrange labour.\nWeek 4-5: Installation commissioning and testing to critical area.\nWeek 6-8: Based on site condition, finalize all installation, provide installation certificate.'
    date_today = datetime.now().strftime('%Y-%m-%d')
    mycursor.execute("INSERT INTO bridge.fee VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                     (quotation,None,1,installation_program,date_today,1,'1-2','3-6','2-4',
                      'Development Approval',0,'Design Development',0,'Tender Documentation',0,
                      'Construction Phase Service',0,None,None,None,None))
    '''bridge.remittances'''
    for i in range(8):
        mycursor.execute("INSERT INTO bridge.remittances VALUES (%s, %s, %s, %s, %s, %s)",
                         (quotation, None, None, None, None, i))
    '''bridge.fee_item'''
    service_list=["Mechanical Service", "CFD Service", "Electrical Service", "Hydraulic Service", "Fire Service","Mechanical Review", "Miscellaneous", "Installation", "Variation"]
    for service in service_list:
        if service=="Variation":
            content_1 = None
            content_2 = None
            content_3 = None
        elif service=="Installation":
            content_1 ='Installation Kickoff'
            content_2 ='Equipment in Position'
            content_3='Installation Completion'
        else:
            content_1=service+' Kickoff'
            content_2=service+' Final Documentation'
            content_3 = None
        mycursor.execute("INSERT INTO bridge.fee_item VALUES (%s, %s, %s, %s, %s, %s)",
                             (quotation, service, content_1, None, None, 0))
        mycursor.execute("INSERT INTO bridge.fee_item VALUES (%s, %s, %s, %s, %s, %s)",
                             (quotation, service, content_2, None, None, 1))
        mycursor.execute("INSERT INTO bridge.fee_item VALUES (%s, %s, %s, %s, %s, %s)",
                             (quotation, service, content_3, None, None, 2))
        mycursor.execute("INSERT INTO bridge.fee_item VALUES (%s, %s, %s, %s, %s, %s)",
                             (quotation, service, None, None, None, 3))

    '''bridge.supplier_fee'''
    service_list = ["Mechanical Service", "CFD Service", "Electrical Service", "Hydraulic Service", "Fire Service",
                    "Mechanical Review", "Miscellaneous", "Installation", "Variation"]
    for service in service_list:
        mycursor.execute("INSERT INTO bridge.supplier_fee VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                             (quotation, service, '',0, None, None, None, None))
    '''bridge.fee_acceptance'''
    mycursor.execute("INSERT INTO bridge.fee_acceptance VALUES (%s, %s, %s, %s)",
                     (quotation, None, None, None))
    '''bridge.emails'''
    date_today=datetime.now().strftime('%Y-%m-%d')
    mycursor.execute("INSERT INTO bridge.emails VALUES (%s, %s, %s, %s, %s, %s, %s)",
                     (quotation, date_today, None, None, None, None, None))
    '''bridge.scope'''
    mycursor.execute("SELECT * FROM bridge.default_scope")
    scope_of_work_database = mycursor.fetchall()
    for i in range(len(scope_of_work_database)):
        minor_or_major=scope_of_work_database[i][0]
        service=scope_of_work_database[i][1]
        scope_type=scope_of_work_database[i][2]
        content=scope_of_work_database[i][3]
        include=scope_of_work_database[i][4]
        row_index=scope_of_work_database[i][5]
        mycursor.execute("INSERT INTO bridge.scope VALUES (%s, %s, %s, %s, %s, %s, %s)",
                         (quotation, minor_or_major, service, scope_type, include, content, row_index))
    '''bridge.major_stage'''
    mycursor.execute("SELECT * FROM bridge.default_stage")
    stage_database=mycursor.fetchall()
    for i in range(len(stage_database)):
        stage=int(stage_database[i][0])+1
        content=stage_database[i][1]
        index=stage_database[i][2]
        mycursor.execute("INSERT INTO bridge.major_stage VALUES (%s, %s, %s, %s, %s)",
                         (quotation, 0, stage, content, index))
    '''bridge.remittance_upload'''
    for i in range(8):
        mycursor.execute("INSERT INTO bridge.remittance_upload VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                     (quotation, 0, 0, 0, 0, 0, 0, i))
    '''bridge.acceptance_upload'''
    mycursor.execute("INSERT INTO bridge.acceptance_upload VALUES (%s, %s, %s, %s, %s, %s, %s)",
                     (quotation, 0, 0, 0, 0, 0, 0))
    '''bridge.supplier_upload'''
    service_list=["Mechanical Service", "CFD Service", "Electrical Service", "Hydraulic Service", "Fire Service","Mechanical Review", "Miscellaneous", "Installation", "Variation"]
    for service in service_list:
        mycursor.execute("INSERT INTO bridge.supplier_upload VALUES (%s, %s, %s, %s, %s, %s)",
                         (quotation,service, 0, 0, 0, 0))
    '''bridge.bill_upload'''
    service_list=["Mechanical Service", "CFD Service", "Electrical Service", "Hydraulic Service", "Fire Service","Mechanical Review", "Miscellaneous", "Installation", "Variation"]
    for service in service_list:
        for i in range(3):
            mycursor.execute("INSERT INTO bridge.bill_upload VALUES (%s, %s, %s, %s)",
                             (quotation, service, 0, i))
    '''bridge.status_colour'''
    mycursor.execute("INSERT INTO bridge.status_colour VALUES (%s, %s, %s, %s, %s, %s, %s)",
                     (quotation, 0, 0, 0, 1,1,1))

    '''bridge.last_update'''
    mycursor.execute("INSERT INTO bridge.last_update VALUES (%s, %s, %s)",
                     (quotation, None, None))

    '''bridge.copilot_sizes'''
    mycursor.execute("INSERT INTO bridge.copilot_sizes VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                     (quotation, 50, 'A3', None, None, None, 50, 'A3', None, None, None))

    '''bridge.copilot_tags'''
    tag_level_list=['Reference Page','','','','','','Basement 2','Basement 1','Ground Level','Level 1',
                    'Level 2','Level 3','Level 4','Level 5','Level 6','Level 7','Level 8','','','','','',
                    '','','','','','Roof']
    for i in range(len(tag_level_list)):
        tag=tag_level_list[i]
        mycursor.execute("INSERT INTO bridge.copilot_tags VALUES (%s, %s, %s, %s)",
                         (quotation, tag, 0, i))

    mydb.commit()
    mycursor.close()
    mydb.close()



def check_or_create_folder(dir):
    if not os.path.exists(dir):
        os.makedirs(dir)

def get_first_name(full_name):
    try:
        return full_name.split(" ")[0]
    except:
        return full_name


def open_pdf_adobe(pdf_dir):
    adobe_address=conf["adobe_address"]
    _thread.start_new_thread(subprocess.call([adobe_address, pdf_dir]), ())


def open_pdf_in_bluebeam(file_name):
    try:
        bluebeam_engine_dir = conf["bluebeam_engine_dir"]
        command= f"Open('{file_name}') View() Close()"
        subprocess.check_output([bluebeam_engine_dir, command])
    except:
        time.sleep(2)
        try:
            bluebeam_engine_dir = conf["bluebeam_engine_dir"]
            command= f"Open('{file_name}') View() Close()"
            subprocess.check_output([bluebeam_engine_dir, command])
        except:
            traceback.print_exc()

def open_sketch(file_dir):
    try:
        _thread.start_new_thread(open_pdf_in_bluebeam, (file_dir, ))
    except:
        traceback.print_exc()



def separate_line(line):
    len_per_line = conf["len_per_line"]
    return textwrap.wrap(line, len_per_line, break_long_words=False)


def char2num(char):
    return ord(char) - ord('A')



def change_quotation_number(current_folder,number):
    try:
        old_folder_name=Path(current_folder).name
        pro_name='-'.join(old_folder_name.split('-')[1:])
        new_folder=os.path.join(Path(current_folder).parent.absolute(),number+'-'+pro_name)
        os.rename(current_folder, new_folder)
        return current_folder, new_folder, True
    except PermissionError:
        return current_folder, new_folder, False


def export_excel_to_pdf(excel_full_path, pdf_path):
    try:
        pythoncom.CoInitialize()
        xlApp = DispatchEx("Excel.Application")
        xlApp.Visible = False
        xlApp.DisplayAlerts = 0
        books = xlApp.Workbooks.Open(excel_full_path, False)
        books.ExportAsFixedFormat(0, pdf_path)
        books.Close(False)
        xlApp.Quit()
        open_pdf_thread(pdf_path)
    except:
        traceback.print_exc()
    finally:
        pythoncom.CoUninitialize()

def export_excel_to_pdf2(excel_full_path, pdf_path,page_list):
    try:
        pythoncom.CoInitialize()
        xlApp = DispatchEx("Excel.Application")
        xlApp.Visible = False
        xlApp.DisplayAlerts = 0
        books = xlApp.Workbooks.Open(excel_full_path, False)
        books.ExportAsFixedFormat(0, pdf_path)
        books.Close(False)
        xlApp.Quit()
        extract_pages(pdf_path, pdf_path, page_list)
        open_pdf_thread(pdf_path)
    except:
        traceback.print_exc()
    finally:
        pythoncom.CoUninitialize()


def export_excel_thread(excel_full_path, pdf_path):
    thread = threading.Thread(target=export_excel_to_pdf, args=(excel_full_path, pdf_path))
    thread.start()

def export_excel_thread2(excel_full_path, pdf_path,page_list):
    thread = threading.Thread(target=export_excel_to_pdf2, args=(excel_full_path, pdf_path,page_list))
    thread.start()


def f_write_designcert_excel(excel_full_path_1,excel_full_path_2,pro_name,drawing_content):
    try:
        workbook = openpyxl.load_workbook(excel_full_path_1)
        worksheet = workbook['Mechanical Design Certificate']
        worksheet.cell(3, 1, pro_name)
        first_cell = None
        for i in range(1, 100):
            if worksheet.cell(row=i, column=2).value == "Drawing Number":
                first_cell = i + 1
                break
        if first_cell is None:
            return False
        if drawing_content != []:
            for i in range(len(drawing_content)):
                drawing_number = drawing_content[i]["Drawing number"]
                drawing_name = drawing_content[i]["Drawing name"]
                drawing_rev = drawing_content[i]["Revision"]
                worksheet.cell(first_cell + i, 2, drawing_number)
                worksheet.cell(first_cell + i, 4, drawing_name)
                worksheet.cell(first_cell + i, 8, drawing_rev)
        workbook.save(excel_full_path_1)

        workbook = openpyxl.load_workbook(excel_full_path_2)
        worksheet = workbook['Mech Design Compliance Cert']
        worksheet.cell(3, 1, pro_name)
        first_cell = None
        for i in range(1, 100):
            if worksheet.cell(row=i, column=2).value == "Drawing Number":
                first_cell = i + 1
                break
        if first_cell is None:
            return False
        if drawing_content != []:
            for i in range(len(drawing_content)):
                drawing_number = drawing_content[i]["Drawing number"]
                drawing_name = drawing_content[i]["Drawing name"]
                drawing_rev = drawing_content[i]["Revision"]
                worksheet.cell(first_cell + i, 2, drawing_number)
                worksheet.cell(first_cell + i, 4, drawing_name)
                worksheet.cell(first_cell + i, 8, drawing_rev)
        workbook.save(excel_full_path_2)
        return True
    except:
        traceback.print_exc()
        return False




def extract_pages(input_pdf, output_pdf, pages_to_extract):
    with open(input_pdf, "rb") as infile:
        reader = PyPDF2.PdfReader(infile)
        writer = PyPDF2.PdfWriter()
        for page_num in pages_to_extract:
            writer.add_page(reader.pages[page_num - 1])
        with open(output_pdf, "wb") as outfile:
            writer.write(outfile)

def f_write_designcert_excel2(excel_full_path_1,pro_name,drawing_content):
    try:
        workbook = openpyxl.load_workbook(excel_full_path_1)
        worksheet = workbook['Mechanical Design Certificate']
        worksheet.cell(3, 1, pro_name)
        first_cell = None
        for i in range(1, 100):
            if worksheet.cell(row=i, column=2).value == "Drawing Number":
                first_cell = i + 1
                break
        if first_cell is None:
            return False
        if drawing_content != []:
            for i in range(len(drawing_content)):
                drawing_number = drawing_content[i]["Drawing number"]
                drawing_name = drawing_content[i]["Drawing name"]
                drawing_rev = drawing_content[i]["Revision"]
                worksheet.cell(first_cell + i, 2, drawing_number)
                worksheet.cell(first_cell + i, 4, drawing_name)
                worksheet.cell(first_cell + i, 8, drawing_rev)

        worksheet = workbook['Mech Design Compliance Cert']
        worksheet.cell(3, 1, pro_name)
        first_cell = None
        for i in range(1, 100):
            if worksheet.cell(row=i, column=2).value == "Drawing Number":
                first_cell = i + 1
                break
        if first_cell is None:
            return False
        if drawing_content != []:
            for i in range(len(drawing_content)):
                drawing_number = drawing_content[i]["Drawing number"]
                drawing_name = drawing_content[i]["Drawing name"]
                drawing_rev = drawing_content[i]["Revision"]
                worksheet.cell(first_cell + i, 2, drawing_number)
                worksheet.cell(first_cell + i, 4, drawing_name)
                worksheet.cell(first_cell + i, 8, drawing_rev)
        workbook.save(excel_full_path_1)
        return True
    except:
        traceback.print_exc()
        return False
def f_write_invoice_excel(excel_dir,info_list,inv_num,pro_name,invoice_info_dict,total_fee,total_inGST):
    try:
        workbook = openpyxl.load_workbook(excel_dir)
        worksheet = workbook['Invoice']
        cur_row = 5
        for i in range(3):
            info_i = info_list[i]
            if info_i != '' and info_i != None:
                worksheet.cell(cur_row, 1, info_i)
                cur_row += 1
        if info_list[3] != None:
            worksheet.cell(cur_row, 1, "ABN\\ACN: " + info_list[3])
        worksheet.cell(4, 10, datetime.today().strftime("%d-%b-%Y"))
        worksheet.cell(5, 10, inv_num)
        worksheet.cell(12, 1, pro_name)

        cur_row = 14
        for service_full,inv_info in invoice_info_dict.items():
            if service_full=='Variation':
                continue
            doc_list=inv_info['doc_list']
            fee_service=inv_info['fee_service']
            worksheet.cell(cur_row, 1, service_full + " design and documentation")
            cur_row += 1
            for i in range(len(doc_list)):
                doc_name=doc_list[i][0]
                fee=doc_list[i][1]
                fee_ingst=doc_list[i][2]
                check=doc_list[i][3]
                worksheet.cell(cur_row, 1, doc_name)
                worksheet.cell(cur_row, 7, float(fee))
                if check:
                    worksheet.cell(cur_row, 8, float(fee))
                    worksheet.cell(cur_row, 9, float(fee))
                    worksheet.cell(cur_row, 10, float(fee_ingst))
                cur_row += 1
            cell=worksheet.cell(cur_row, 6, "Total: ")
            cell.font = Font(bold=True)
            cell=worksheet.cell(cur_row, 7, fee_service)
            cell.font = Font(bold=True)
            cur_row += 2
        cur_row += 1

        if 'Variation' in list(invoice_info_dict.keys()):
            doc_list=invoice_info_dict['Variation']['doc_list']
            worksheet.cell(cur_row, 1, "Variation")
            cur_row += 1
            for i in range(len(doc_list)):
                doc_name=doc_list[i][0]
                fee=doc_list[i][1]
                fee_ingst=doc_list[i][2]
                check=doc_list[i][3]
                worksheet.cell(cur_row, 1, doc_name)
                worksheet.cell(cur_row, 7, float(fee))
                if check:
                    worksheet.cell(cur_row, 8, float(fee))
                    worksheet.cell(cur_row, 9, float(fee))
                    worksheet.cell(cur_row, 10, float(fee_ingst))
                cur_row += 1
        worksheet.cell(44, 9, inv_num)
        worksheet.cell(39, 9, float(total_fee))
        worksheet.cell(39, 10, float(total_inGST))
        workbook.save(excel_dir)
        return True
    except PermissionError:
        traceback.print_exc()
        return "Please close the preview or file before you use it"
    except:
        traceback.print_exc()
        return "Some errors happen."


def format_list_with_and(content_list):
    if len(content_list) == 0:
        return ""
    elif len(content_list) == 1:
        return content_list[0]
    elif len(content_list) == 2:
        return " and ".join(content_list)
    else:
        return ", ".join(content_list[:-1]) + ", and " + content_list[-1]




def f_write_minor_fee_excel(excel_dir,full_name,company,address,first_name,quo_num,email_fee_proposal_date,fee_rev,pro_name,
                            time_fee_proposal_Fee_proposal,time_fee_proposal_Pre_design,time_fee_proposal_Documentation,
                            service_type_list,services_content_dict,past_projects,total_fee,total_ist):
    row_per_page = 46
    try:
        '''Page 1'''
        workbook=openpyxl.load_workbook(excel_dir)
        worksheet=workbook['Fee Proposal']
        worksheet.cell(2, 2, full_name)
        worksheet.cell(3, 2, company)
        worksheet.cell(4, 2, address)
        worksheet.cell(5, 1, f'Dear {first_name},')
        worksheet.cell(2, 8, quo_num)
        worksheet.cell(3, 8, email_fee_proposal_date)
        worksheet.cell(4, 8, fee_rev)
        worksheet.cell(6, 1, "Re: " + pro_name)
        worksheet.cell(8,1, f"Thank you for giving us the opportunity to submit this fee proposal for our {format_list_with_and(service_type_list)} for the above project.")
        worksheet.cell(16, 7, time_fee_proposal_Fee_proposal)
        worksheet.cell(21, 7, time_fee_proposal_Pre_design)
        worksheet.cell(26, 7, time_fee_proposal_Documentation)
        '''Page 2~6'''
        cur_index = 0
        for i in range(len(service_type_list)):
            service_type=service_type_list[i]
            service_info=services_content_dict[service_type]
            if i==0:
                cur_row = 52
            else:
                cur_row = 84 + (i - 1) * row_per_page
            content_type_list=['Extent', 'Clarifications', 'Deliverables']
            for content_type in content_type_list:
                content_list=service_info[content_type]
                if content_list != []:
                    cur_index += 1
                    cell = worksheet.cell(cur_row, 1, "2." + str(cur_index))
                    cell.font = Font(bold=True)
                    cell = worksheet.cell(cur_row, 2, service_type + "-" + content_type)
                    cell.font = Font(bold=True)
                    cur_row += 1
                    if content_type=='Clarifications':
                        worksheet.cell(cur_row, 2, "Our fee does not yet include:")
                        worksheet.cell(cur_row, 2).font = Font(bold=False)
                        worksheet.cell(cur_row, 2).alignment = Alignment(horizontal='left')
                        cur_row += 1
                    for content_i in content_list:
                        worksheet.cell(cur_row, 1, "•")
                        for line in separate_line(content_i):
                            worksheet.cell(cur_row, 2, line)
                            cur_row += 1
                    cur_row += 1
        '''Page 7'''
        cur_row = 284
        for project in past_projects:
            worksheet.cell(cur_row, 1, "•")
            for line in separate_line(project):
                worksheet.cell(cur_row, 2, line)
                cur_row += 1

        cur_row = 320
        for i in range(len(service_type_list)):
            service_type=service_type_list[i]
            service_info=services_content_dict[service_type]
            fee=service_info['Fee'][0]
            fee_ingst=service_info['Fee'][1]
            worksheet.cell(cur_row + i, 2, service_type + " design and documentation")
            worksheet.cell(cur_row + i, 6, float(fee))
            worksheet.cell(cur_row + i, 7, float(fee_ingst))
        # cur_row = 325
        # worksheet.cell(cur_row, 6, float(total_fee))
        # worksheet.cell(cur_row, 6, float(total_ist))
        cur_row = 338
        worksheet.cell(cur_row,2, "Re: " + pro_name)

        '''Delete rows'''
        if len(service_type_list)==1:
            cur_row = 325
            worksheet.cell(cur_row, 3, '')
            worksheet.cell(cur_row, 6, '')
            worksheet.cell(cur_row, 7, '')

        fee_table_delete_dict={1:[321,322,323,324],2:[322,323,324],3:[323,324],4:[324],5:[]}
        fee_table_delete=fee_table_delete_dict[len(service_type_list)]
        for row in fee_table_delete:
            worksheet.row_dimensions[row].hidden = True
        worksheet.row_dimensions[353].height = int((1.16+0.42*len(fee_table_delete))/0.035)

        page_num_row_list=[42]+[82,128,174,220,266][:len(service_type_list)]+[307,353,392,439]
        for i in range(len(page_num_row_list)):
            total_page=str(len(page_num_row_list))
            page_i=str(i+1)
            page_name='Page '+page_i+' of '+total_page
            worksheet.cell(page_num_row_list[i], 8, page_name)

        print_area_dict={1:'A1:I82,A267:I439',2:'A1:I128,A267:I439',3:'A1:I174,A267:I439',4:'A1:I220,A267:I439',5:'A1:I439',}
        print_area=print_area_dict[len(service_type_list)]
        worksheet.print_area=print_area

        workbook.save(excel_dir)
        return True
    except PermissionError:
        traceback.print_exc()
        return "Please close the preview or file before you use it"
    except FileNotFoundError:
        traceback.print_exc()
        return "Please Contact Administer, the app cant find the file"
    except:
        traceback.print_exc()
        return 'Some errors happen'



def f_write_major_fee_excel(excel_dir,pro_name,service_name,full_name,company,address,email_fee_proposal_date,quo_num,
                            first_name,service_type_list,pro_note,stage_content_dict,services_content_dict,past_projects,
                            stage_name_list,stage_fee_dict,fee_rev):
    row_per_page=46
    try:
        '''Page 1'''
        workbook=openpyxl.load_workbook(excel_dir)
        worksheet=workbook['Fee Proposal']
        worksheet.cell(27, 2, pro_name)
        worksheet.cell(31, 2, service_name + " Service Fee Proposal")
        worksheet.cell(36, 1, full_name)
        worksheet.cell(37, 1, company)
        worksheet.cell(38, 1, address)
        worksheet.cell(41, 1, email_fee_proposal_date)
        worksheet.cell(41, 7, quo_num)
        '''Page 2'''
        worksheet.cell(50, 1, f"Dear {first_name},")
        worksheet.cell(51, 1, "Re: " + pro_name)
        worksheet.cell(53,1, f"Thank you for giving us the opportunity to submit this fee proposal for our {format_list_with_and(service_type_list)} for the above project.")
        '''Page 3'''
        worksheet.cell(86,2, "We have prepared this submission in response to an invitation of " + full_name + " for the provision of our consulting services for " +pro_name+'.')
        cur_row = 97
        if pro_note!='':
            for line in pro_note.split("\n"):
                for line_i in separate_line(line):
                    worksheet.cell(cur_row, 2, line_i)
                    cur_row += 1
        cur_row += 1
        cell=worksheet.cell(cur_row, 1, "2.3")
        cell.font = Font(bold=True)
        cell=worksheet.cell(cur_row, 2, "Deliverables at each stage")
        cell.font = Font(bold=True)
        cur_row += 1
        for stage in stage_name_list:
            content_list=stage_content_dict[stage]
            # print([cur_row,2,stage])
            cell = worksheet.cell(cur_row, 2, stage)

            cell.font = Font(bold=True)
            cur_row += 1
            print(cur_row)
            for content in content_list:
                worksheet.cell(cur_row, 1, "•")
                worksheet.cell(cur_row, 2, content)
                cur_row += 1
            # cur_row += 1
        '''Page 4~8'''
        cur_index = 0
        for i in range(len(service_type_list)):
            cur_row = 131 + i * row_per_page
            if i==0:
                cur_row+=1
            service_type=service_type_list[i]
            service_info=services_content_dict[service_type]
            content_type_list=['Extent', 'Clarifications', 'Deliverables']
            for content_type in content_type_list:
                content_list=service_info[content_type]
                if content_list != []:
                    cur_index+=1
                    cell = worksheet.cell(cur_row, 1, "3." + str(cur_index))
                    cell.font = Font(bold=True)
                    cell = worksheet.cell(cur_row, 2, service_type + "-" + content_type)
                    cell.font = Font(bold=True)
                    cur_row += 1
                    if content_type=='Clarifications':
                        worksheet.cell(cur_row, 2, "Our fee does not yet include:")
                        worksheet.cell(cur_row, 2).font = Font(bold=False)
                        worksheet.cell(cur_row, 2).alignment = Alignment(horizontal='left')
                        cur_row += 1
                    for content_i in content_list:
                        worksheet.cell(cur_row, 1, "•")
                        for line in separate_line(content_i):
                            worksheet.cell(cur_row, 2, line)
                            cur_row += 1
                    cur_row += 1
        '''Page 9'''
        cur_row=376
        for project in past_projects:
            worksheet.cell(cur_row, 1, "•")
            for line in separate_line(project):
                worksheet.cell(cur_row, 2, line)
                cur_row += 1

        '''Page 10'''
        n_stage=len(stage_name_list)
        merge_list=['412','413','414','415','416','417','418','419']
        if n_stage==1:
            for merge_row in merge_list:
                merge_content='D'+merge_row+':'+'G'+merge_row
                worksheet.merge_cells(merge_content)
            merge_content_list = ['D410:G411', 'H410:H411']
            for merge_content in merge_content_list:
                worksheet.merge_cells(merge_content)
            cell = worksheet.cell(410, 8, 'Total')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif n_stage==2:
            for merge_row in merge_list:
                merge_content='D'+merge_row+':'+'E'+merge_row
                worksheet.merge_cells(merge_content)
                merge_content = 'F' + merge_row + ':' + 'G' + merge_row
                worksheet.merge_cells(merge_content)
            merge_content_list = ['D410:E411', 'F410:G411', 'H410:H411']
            for merge_content in merge_content_list:
                worksheet.merge_cells(merge_content)
            cell = worksheet.cell(410, 8, 'Total')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif n_stage==3:
            for merge_row in merge_list:
                merge_content = 'G' + merge_row + ':' + 'H' + merge_row
                worksheet.merge_cells(merge_content)
                cell_sum = 'G' + merge_row
                cell_function = '=D' + merge_row + '+E' + merge_row + '+F' + merge_row
                worksheet[cell_sum] = cell_function
            merge_content_list = ['D410:D411', 'E410:E411', 'F410:F411', 'G410:H411']
            for merge_content in merge_content_list:
                worksheet.merge_cells(merge_content)
            cell = worksheet.cell(410, 7, 'Total')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            merge_content_list = ['D410:D411', 'E410:E411', 'F410:F411', 'G410:G411', 'H410:H411']
            for merge_content in merge_content_list:
                worksheet.merge_cells(merge_content)
            cell = worksheet.cell(410, 8, 'Total')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        column_dict={1:[4],2:[4,6],3:[4,5,6],4:[4,5,6,7]}
        column_list=column_dict[n_stage]
        for i in range(len(column_list)):
            column=column_list[i]
            stage_name=stage_name_list[i]
            cell = worksheet.cell(410, column, stage_name)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for i in range(len(service_type_list)):
            row=412+i
            service_name=service_type_list[i]
            worksheet.cell(row, 2, service_name)
            stage_fee_list=stage_fee_dict[service_name]
            for j in range(len(stage_fee_list)):
                column=column_list[j]
                worksheet.cell(row, column, float(stage_fee_list[j]))
        worksheet.cell(435, 2, "Re: " + pro_name)

        '''Delete rows'''
        fee_table_delete_dict={1:[413,414,415,416],2:[414,415,416],3:[415,416],4:[416],5:[]}
        fee_table_delete=fee_table_delete_dict[len(service_type_list)]
        if n_stage==1 and len(service_type_list)==1:
            fee_table_delete.append(417)
        for row in fee_table_delete:
            worksheet.row_dimensions[row].hidden = True
        worksheet.row_dimensions[450].height = int((1.19+0.42*len(fee_table_delete))/0.035)

        page_num_row_list=[83,129]+[175,221,267,313,359][:len(service_type_list)]+[404,450,480,520]
        for i in range(len(page_num_row_list)):
            total_page=str(len(page_num_row_list))
            page_i=str(i+1)
            page_name='Page '+page_i+' of '+total_page
            worksheet.cell(page_num_row_list[i], 8, page_name)
            quo_rev=quo_num+' '+'Rev '+fee_rev
            worksheet.cell(page_num_row_list[i], 5, quo_rev)

        print_area_dict={1:'A1:I175,A360:I520',2:'A1:I221,A360:I520',3:'A1:I267,A360:I520',4:'A1:I313,A360:I520',5:'A1:I520',}
        print_area=print_area_dict[len(service_type_list)]
        worksheet.print_area=print_area

        workbook.save(excel_dir)
        return True

    except PermissionError:
        traceback.print_exc()
        return "Please close the preview or file before you use it"
    except FileNotFoundError:
        traceback.print_exc()
        return "Please Contact Administer, the app cant find the file"
    except:
        traceback.print_exc()
        return "Some errors happen"




def f_write_installation_fee_excel(excel_dir,full_name,company,address,quo_num,install_fee_date,
                                   fee_installation_rev,pro_name,drawing_content,content_dict,
                                   past_projects,program,installation_content_list):
    try:
        '''Page 1'''
        workbook = openpyxl.load_workbook(excel_dir)
        worksheet = workbook['Fee Proposal']
        first_name = get_first_name(full_name)
        worksheet.cell(2, 2, full_name)
        worksheet.cell(3, 2, company)
        worksheet.cell(4, 2, address)
        worksheet.cell(2, 8, quo_num)
        worksheet.cell(3, 8, install_fee_date)
        worksheet.cell(4, 8, fee_installation_rev)
        worksheet.cell(5, 1, f'Dear {first_name},')
        worksheet.cell(6, 1, "Re: " + pro_name)
        '''Page 2'''
        cur_row = 51
        for i in range(len(drawing_content)):
            worksheet.cell(cur_row + i, 2, drawing_content[i]["Drawing number"])
            worksheet.cell(cur_row + i, 4, drawing_content[i]["Drawing name"])
            worksheet.cell(cur_row + i, 8, drawing_content[i]["Revision"])
        '''Page 3'''
        cur_row = 94
        content_type_list = ['Extent', 'Clarifications', 'Deliverables']
        title_name_dict = {"Extent": "Mechanical Service Installation-Extent ",
                           "Clarifications": "Mechanical Service Installation-Clarifications ",
                           "Deliverables": "Mechanical Equipment List", }
        subpoint=0
        for content_type in content_type_list:
            title_name = title_name_dict[content_type]
            content_list=content_dict[content_type]
            if content_list != []:
                cell=worksheet.cell(cur_row, 1, f"2.{subpoint + 1}")
                subpoint=subpoint+1
                cell.font = Font(bold=True)
                cell=worksheet.cell(cur_row, 2, title_name)
                cell.font = Font(bold=True)
                if subpoint==2:
                    cur_row += 1
                    worksheet.cell(cur_row, 2, 'Our fee does not yet include:')
                cur_row += 1
                for content in content_list:
                    worksheet.cell(cur_row, 1, "•")
                    for line in separate_line(content):
                        worksheet.cell(cur_row, 2, line)
                        cur_row += 1
                cur_row += 1
        '''Page 4'''
        cur_row = 164
        for i, project in enumerate(past_projects):
            worksheet.cell(cur_row + i, 1, "•")
            worksheet.cell(cur_row + i, 2, project)
        cur_row = 177
        if program != '':
            for line in program.split("\n"):
                worksheet.cell(cur_row, 1, "•")
                worksheet.cell(cur_row, 2, line)
                cur_row += 1
        '''Page 5'''
        cur_row = 195
        if len(installation_content_list) > 0:
            for i in range(len(installation_content_list)):
                worksheet.cell(cur_row + i, 2, installation_content_list[i][0])
                worksheet.cell(cur_row + i, 6, int(float(installation_content_list[i][1])))
                # worksheet.cell(cur_row + i, 7, installation_content_list[i][2])
        cur_row = 214
        worksheet.cell(cur_row, 2, "Re: " + pro_name)


        '''Delete rows'''
        fee_table_delete_dict={1:[196,197,198],2:[197,198],3:[198],4:[]}
        fee_table_delete=fee_table_delete_dict[len(installation_content_list)]
        for row in fee_table_delete:
            worksheet.row_dimensions[row].hidden = True
        worksheet.row_dimensions[230].height = int((2.38+0.53*len(fee_table_delete))/0.035)
        workbook.save(excel_dir)
        return True
    except PermissionError:
        traceback.print_exc()
        return "Please close the preview or file before you use it"
    except FileNotFoundError:
        traceback.print_exc()
        return "Please Contact Administer, the app cant find the file"
    except:
        traceback.print_exc()
        return "Some errors happen."



def is_valid_email(email):
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None


def convert_date(date_str):
    for fmt in ["%d-%b-%Y", "%Y-%m-%d"]:
        try:
            return datetime.strptime(date_str, fmt).strftime("%d %B %Y")
        except:
            return date_str



def f_get_client_xero_id(client_id):
    try:
        mydb = mysql.connector.connect(host="192.168.1.199", user="bridge", password="PcE$yD2023", database="bridge")
        mycursor = mydb.cursor()
        mycursor.execute("SELECT * FROM bridge.clients WHERE client_id = %s", (client_id,))
        client_database=mycursor.fetchall()
        if len(client_database)==1:
            xero_client_id=client_database[0][1]
        else:
            xero_client_id=None
        return xero_client_id
    except:
        traceback.print_exc()
    finally:
        mycursor.close()
        mydb.close()


def parse_when_where(content):
    when_pattern = r"When:\s*(.+)"
    where_pattern = r"Where:\s*(.+)"
    when_match = re.search(when_pattern, content)
    when_info = when_match.group(1) if when_match else "Not found"
    where_match = re.search(where_pattern, content)
    where_info = where_match.group(1) if where_match else "Not found"
    return when_info, where_info


def check_excel_open(file_path):
    try:
        temp_path = file_path + ".temp"
        os.rename(file_path, temp_path)
        os.rename(temp_path, file_path)
        return False
    except:
        return True

def create_folder(file_path):
    if not os.path.exists(file_path):
        os.mkdir(file_path)